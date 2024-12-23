from openpyxl import load_workbook,Workbook
import numpy as np
import os,sys
from tkinter import *
from tkcalendar import *
from tkdbprocess import *

#to access the path of file
def resource(relative_path):
     try:
          base_path=sys._MEIPASS
     except Exception:
           base_path=os.path.abspath(".")
     return os.path.join(base_path,relative_path)



#load or create excell file
try:
    wb=load_workbook(resource("src/new.xlsx"))
    ws=wb.active
    if ws["A1"].value=="loan_date":
        pass
    else:
        head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
        ws.append(head)
        wb.save(resource("src/new.xlsx"))
    
except:
    wb=Workbook()
    ws=wb.active
    head=["loan_date","bill_no","name","co_name","street","address","int_amt","weight","item","no_item","Phone No","release"]
    ws.append(head)
    wb.save(resource("src/new.xlsx")) 


#for save new data to excell
def ws_save(loan_data):
    ws.append(list(loan_data.values()))
    wb.save(resource("src/new.xlsx"))
    
    return True

#update the release date of released loan bill number 
def xlupdate_release(bill_no,release_date):
    id=find_bill(int(bill_no))
    ws[f"l{id+1}"]=release_date
    wb.save(resource("src/new.xlsx"))

#find max bill number
def max_bill():
    s=list(ws.columns)[1]
    c=[]
    for i in range(1,ws.max_row):
        try:
            a=str(s[i].value)
            c.append(int(a))
        except:
            pass
    c.sort()
    return c[-1]

#find row number based on bill number for update/delete task
def find_bill(b_no):
    s=list(ws.columns)[1]
    c=[]
    for i in range(1,ws.max_row):
        try:
            if b_no==int(str(s[i].value)):
                return i
        except:
            pass
    else:
        return 0

#alter the value 
def alter(id,uplist):
        id=find_bill(id)
        id+=1
        ws[f"C{id}"]=uplist[2]    #name
        ws[f"D{id}"]=uplist[3]    #coname
        ws[f"A{id}"]=uplist[0]    #loandate
        ws[f"B{id}"]=uplist[1]    #billno
        ws[f"F{id}"]=uplist[5]    #address
        ws[f"H{id}"]=uplist[7]    #weight
        ws[f"I{id}"]=uplist[8]    #items
        ws[f"G{id}"]=uplist[6]    #intial amount
        ws[f"E{id}"]=uplist[4]    #street
        ws[f"J{id}"]=uplist[9]    #no item
        ws[f"k{id}"]=uplist[10]   #ph no
        ws[f"l{id}"]=uplist[11]   #release date
        
        wb.save(resource("src/new.xlsx"))

    
#delete record from excell base on bill number
def delete(id):
    id=find_bill(id)
    if id:
        ws.delete_rows((int(id)+1))
        wb.save(resource("src/new.xlsx"))
    else:
        return False
    

    

#sample funtion for search based on excell data only (not used  in this).
def srch(i,val):
    
        id=val
        i=str(i)
        if i=="கடன் தேதி":
            src='loan_date'
        elif i=="கடன் எண்":
            src='bill_no'
        elif i=="பெயர்":
            src='name'
        elif i=="த/க பெயர்":
            src='co_name'
        elif i=="ஊர்":
            src='address'
        elif i=="கடன் தொகை":
            src='amount'
        elif i=="பொருள்":
            src='items'
        elif i=="மீட்ட தேதி":
            src='relese_date'
        elif i=="எடை":
            src='weight'
        elif i=='Phone No':
            src="Phone No"
        else:
            src="none"
        

        
        op=[]
        c=False
        x=[]
        if src=='loan_date':
            id=datechange(id)
            try:
                s=list(ws.columns)[0]
                for i in range(1,(ws.max_row)):
                    if datechange(s[i].value)==(id):
                        op.append(i)
                    else:
                        raise ValueError

            except ValueError:
                c=True
                
        
    
        elif src=='bill_no':
            try:
                s=list(ws.columns)[1]
                for i in range(1,ws.max_row):
                    a=str(s[i].value)
                    if int(a)==int(id):
                        op.append(i)
                    else:
                        raise ValueError

            except ValueError:        
                c=True


        elif src=='name':
            s=np.array(list(ws.columns)[2])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='co_name':
            s=list(ws.columns)[3]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='address':
            s=np.array(list(ws.columns)[5])
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='amount':
            s=list(ws.columns)[6]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass

        elif src=='weight':
            s=list(ws.columns)[7]
            for i in range(1,ws.max_row):
                a=str(s[i].value)
                try:
                    if int(a)==int(id):
                        op.append(i)
                except:
                    pass
        
        elif src=='items':
            s=list(ws.columns)[8]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='Phone No':
            s=list(ws.columns)[10]
            for i in range(1,ws.max_row):
                if str(id) in str(s[i].value):
                    op.append(i)

        elif src=='relese_date':
            id=datechange(id)
            s=list(ws.columns)[11]
            for i in range(1,ws.max_row):
                if datechange(s[i].value)==datechange(id):
                    op.append(i)

        if c==False:
            for a in op:
                c=list(ws.rows)[int(a)]
                v=[]
                for i in range(0,12):
                    # if c[i].value==None:
                    #     pass
                    v.append(c[i].value)
                v.append(a)
                x.append(v)
        
        return x
